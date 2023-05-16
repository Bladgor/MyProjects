from openpyxl import load_workbook

data = {1: 11.5, 2: 11.5, 3: 16, 4: 16, 5: 20.5, 6: 20.5, 7: 25, 8: 25, 9: 29.5,
        10: 29.5, 11: 34, 12: 34, 13: 38.5, 14: 38.5, 15: 43, 16: 43}

up_down = {1: 0, 2: 25, 3: 37, 4: 49, 5: 62, 6: 77}


def check(num):
    if num % 1 == 0:
        return f'{round(num)}'
    else:
        return f'{round(num // 1)},{round((num % 1) * 10)}'


wb = load_workbook('Registers.RegistersBrw.xlsx')
ws = wb['List1']
current_cell = (ws['B2']).value  # Значение первой ячейки
index = 2
while current_cell:
    ws[f'C{index}'] = check(data[int(current_cell[1:3])] + ((int(current_cell[4:6]) - 1) * 0.7) +
                            up_down[(int(current_cell[7:]))])
    index += 1
    current_cell = (ws[f'B{index}']).value

wb.save('file_штабик.xlsx')
