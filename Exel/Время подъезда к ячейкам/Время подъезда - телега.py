from openpyxl import load_workbook

data = {1: 4.5, 2: 5, 3: 8.7, 4: 9.4, 5: 13.1, 6: 13.8, 7: 17.9, 8: 18.6, 9: 22.3,
        10: 23, 11: 26.7, 12: 27.4, 13: 31.5, 14: 32.2, 15: 35.9, 16: 36.6}


def check(num):
    if num % 1 == 0:
        return f'{round(num)}'
    else:
        return f'{round(num // 1)},{round((num % 1) * 10)}'


wb = load_workbook('Registers.RegistersBrw.xlsx')
ws = wb['List1']
current_cell = (ws['B2']).value  # Значение первой ячейки
index = 2
while current_cell:
    if current_cell[4:6] == '01':
        ws[f'C{index}'] = check(data[int(current_cell[1:3])])
    else:
        ws[f'C{index}'] = check(data[int(current_cell[1:3])] + ((int(current_cell[4:6]) - 1) * 0.7))
    index += 1
    current_cell = (ws[f'B{index}']).value

wb.save('file.xlsx')
