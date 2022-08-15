# import openpyxl
from openpyxl import load_workbook


def check_value(key, value, check_dict, error_dict):
    """
    Функция проверяет одиночные значения, не было ли перемещения на эту ячейку с двух
    других ячеек.
    Возвращает словарь со всеми такими ячейками.
    """
    for elem in check_dict:
        if elem != key:
            if value in check_dict[elem]:
                if value in error_dict:
                    error_dict[value].append(elem)
                    if key not in error_dict:
                        error_dict[value].append(key)
                else:
                    error_dict[value] = [elem]
    return error_dict


# Load in the workbook
wb = load_workbook('file_1.xlsx')
ws = wb['Лист1']  # В каком листе проверяем
my_dict = dict()  # Словарь со всеми значениями из какой ячейки в какую переместили
dict_error = dict()  # Словарь, в который будут добавляться ячейки, из которых товар был перемещён на несколько других
dict_one = dict()

current_cell = (ws['A1']).value  # Значение первой ячейки
index = 1
while current_cell is not None:  # Пока ячейка не пуста
    if current_cell in my_dict:  # Если ячейка есть в словаре
        if (ws[f'B{index}']).value not in my_dict[current_cell]:  # Если в значении (списке) нет
            my_dict[current_cell].append((ws[f'B{index}']).value)  # то добавляем
    else:
        my_dict[current_cell] = [(ws[f'B{index}']).value]  # Иначе создаём ключ со значением
    index += 1
    current_cell = (ws[f'A{index}']).value  # Переходим к следующей ячейке

wb.save('file.xlsx')

for elem in my_dict:  # Здесь находим все списки, в которых больше одного значения (ячейки)
    if len(my_dict[elem]) > 1:  # И если есть, то добавляем в словарь ошибок
        dict_error[elem] = my_dict[elem]

for elem in my_dict:  # Пробегаемся по всем одиночным значениям
    current_value = my_dict[elem][0]
    dict_one = check_value(elem, current_value, my_dict, dict_one)

for elem in dict_error:
    print(f'Откуда {elem}: -> Куда {dict_error[elem]}\t')
print('=' * 40)
for elem in dict_one:
    print(f'Откуда {dict_one[elem]}: -> \tКуда {elem}\t')
