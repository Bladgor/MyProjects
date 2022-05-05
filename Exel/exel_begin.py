from datetime import datetime
import openpyxl
import pprint

wb = openpyxl.load_workbook(filename='./ЛЗ_май.xlsx')
sheet = wb['заказы']
print(wb.sheetnames)

sheet['B2'] = 6
print(sheet['B2'].value)

row_4 = [x.value for x in sheet[4]]

index_date_calculate = row_4.index('Дата-просчитан') + 1
cell_calculate = sheet.cell(row=4, column=index_date_calculate)
date_calculate_column = cell_calculate.column_letter

index_date_selected = row_4.index('Дата-подобран') + 1
cell_selected = sheet.cell(row=4, column=index_date_selected)
date_selected_column = cell_selected.column_letter

date_format = '%d%m%Y'
my_dict = {x.value: {'Дата-просчитан': type(sheet[f'{date_calculate_column}{x.row}'].value),
                     'Дата-подобран': str(sheet[f'{date_selected_column}{x.row}'].value)} for x in sheet['A']}
print(my_dict)

# wb.save('ЛЗ_май.xlsx')
