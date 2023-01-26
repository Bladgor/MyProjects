from openpyxl import load_workbook
from search_column import search_column
from pprint import pprint

wb = load_workbook('Registers.RegistersBrw.xlsx')  # Загружаем файл
ws = wb['List1']  # В каком листе проверяем

party_column = search_column(ws, 'Партия')
q_column = search_column(ws, 'Q общ. баз.')
cell_address_column = search_column(ws, 'Адрес ячейки')
product_column = search_column(ws, 'Продукт')
description_column = search_column(ws, 'Наименование')

dict_party = dict()  # Здесь будут все партии с их количеством

index = 5

party_cell = (ws[f'{party_column}{index}']).value
q_cell = (ws[f'{q_column}{index}']).value
cell_address_cell = (ws[f'{cell_address_column}{index}']).value
product_cell = (ws[f'{product_column}{index}']).value
description_cell = (ws[f'{description_column}{index}']).value

while party_cell:
    if party_cell in dict_party:
        dict_party[party_cell]['quantity'] += 1
        dict_party[party_cell]['total'] += q_cell
        dict_party[party_cell]['cell_addresses'][cell_address_cell] = q_cell
    else:
        dict_party[party_cell] = {
            'quantity': 1,
            'total': q_cell,
            'product': product_cell,
            'description': description_cell,
            'cell_addresses': {cell_address_cell: q_cell},
        }
    index += 1
    party_cell = (ws[f'{party_column}{index}']).value
    q_cell = (ws[f'{q_column}{index}']).value
    cell_address_cell = (ws[f'{cell_address_column}{index}']).value
    product_cell = (ws[f'{product_column}{index}']).value
    description_cell = (ws[f'{description_column}{index}']).value

pprint(dict_party)

wb.create_sheet('Лист1', 0)
ws_1 = wb['Лист1']
max_total = 335

print(wb.active)

index_new = 1
for elem in dict_party:
    quantity = dict_party[elem]['quantity']
    total = dict_party[elem]['total']

    if quantity > 1 and total <= max_total:
        print(index_new)

        cell_address_cell = (ws_1[f'A{index_new}'])
        product_cell = (ws_1[f'B{index_new}'])
        description_cell = (ws_1[f'C{index_new}'])
        q_cell = (ws_1[f'D{index_new}'])
        party_cell = (ws_1[f'E{index_new}'])
        index_new += 1

        for addresses in dict_party[elem]['cell_addresses']:
            cell_address_cell.value = addresses
            q_cell.value = dict_party[elem]['cell_addresses'][addresses]
            product_cell.value = dict_party[elem]['product']
            description_cell.value = dict_party[elem]['description']
            party_cell.value = elem



print(index_new)
wb.save('file.xlsx')
