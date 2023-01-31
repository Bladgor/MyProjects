from openpyxl import load_workbook
from search_column import search_column

quant_mono_dict = {'100191': 480, '112267': 480, '127350': 480, '132157': 480, '100194': 288, '100544': 168,
                   '100545': 552, '102164': 33, '102182': 33, '103694': 3360, '103915': 624, '103916': 5280,
                   '104535': 33, '105742': 33, '105850': 480, '106023': 33, '106209': 33, '107686': 336, '107980': 336,
                   '109235': 624, '109283': 552, '109342': 552, '109372': 168, '109391': 5280, '109624': 33,
                   '110612': 624, '110669': 2880, '110670': 2880, '110671': 2880, '110672': 2880, '110673': 2880,
                   '110675': 2880, '110676': 2880, '110681': 2880, '110682': 2880, '110683': 2880, '110684': 2880,
                   '110685': 2880, '111158': 33, '111159': 33, '111320': 33, '111321': 33, '111322': 33, '111323': 33,
                   '111324': 33, '111325': 33, '111326': 33, '111327': 33, '111328': 33, '111329': 33, '111330': 33,
                   '111331': 33, '111332': 33, '111333': 33, '111334': 33, '111335': 33, '111336': 33, '111337': 33,
                   '111338': 33, '111339': 33, '111340': 33, '111341': 33, '111342': 33, '111343': 33, '111344': 33,
                   '111581': 33, '111701': 2880, '111702': 2880, '111703': 2880, '111704': 2880, '111705': 2880,
                   '111706': 2880, '111707': 2880, '111708': 2880, '111709': 2880, '111710': 2880, '111711': 2880,
                   '111712': 2880, '111714': 2880, '111716': 2880, '111718': 2880, '111720': 2880, '112119': 33,
                   '112120': 12, '112231': 33, '112232': 33, '113193': 552, '113403': 168, '113404': 552, '113669': 12,
                   '113772': 33, '115196': 336, '115209': 33, '115615': 33, '115728': 2880, '120681': 624,
                   '120757': 264, '121574': 552, '121576': 480, '122367': 33, '122487': 60, '122667': 60, '123318': 12,
                   '124124': 384, '124398': 552, '124444': 264, '124820': 216, '125875': 6, '126314': 1080,
                   '126392': 540, '126650': 264, '126651': 552, '126863': 60, '127940': 20, '128642': 234, '128647': 60,
                   '128648': 60, '128649': 60, '128908': 552, '129062': 6, '129450': 20, '129934': 12, '129935': 12,
                   '129936': 12, '129937': 12, '129965': 6, '130430': 8, '130779': 60, '131334': 144, '131460': 60,
                   '131892': 12, '131896': 552, '131897': 528, '131908': 12, '131917': 552, '131968': 32, '132058': 8,
                   '132069': 8, '132070': 8, '132071': 8, '132072': 8, '132126': 552, '132228': 552, '132229': 552,
                   '132231': 552, '132237': 120, '132238': 120, '132247': 528, '132284': 16, '132292': 528,
                   '132309': 690, '132316': 117, '132317': 117, '132328': 0, '132330': 480, '132343': 0, '132344': 0,
                   '132347': 144, '132348': 552, '132350': 5280, '132375': 144, '132376': 144, '132378': 4, '132379': 6,
                   '132381': 552, '132410': 0, '132411': 480, '132416': 12, '132457': 144, '132460': 63, '132471': 264,
                   '132488': 480, '132495': 234, '132503': 552, '132517': 480, '132518': 144, '132538': 12,
                   '132541': 312, '132543': 312, '132544': 312, '132545': 552, '132583': 312, '132593': 144,
                   '132640': 552, '132662': 16, '132663': 16, '132666': 480, '132713': 24, '132714': 18, '132715': 24,
                   '132717': 12, '132737': 408, '132738': 48, '132772': 4, '132773': 690, '132812': 144, '132813': 312,
                   '132837': 117, '132838': 552, '132874': 12, '308681': 33, '436802': 60, '100548': 750, '100668': 864,
                   '100669': 12240, '100670': 840, '100671': 6000, '100674': 480, '101202': 720, '102522': 63,
                   '103932': 432, '103933': 480, '104384': 576, '104686': 600, '105865': 600, '106337': 864,
                   '107526': 750, '108271': 432, '108345': 360, '108455': 540, '124320': 750, '124441': 840,
                   '125799': 600, '126756': 750, '127118': 750, '127903': 312, '128082': 750, '128833': 840,
                   '129027': 456, '129028': 240, '129063': 480, '129218': 63, '129895': 750, '129939': 384,
                   '130059': 750, '130895': 750, '130900': 750, '131066': 750, '131269': 540, '131301': 750,
                   '131592': 540, '131662': 6120, '131663': 6120, '131664': 6120, '131665': 6120, '131703': 312,
                   '131741': 612, '131742': 750, '131743': 750, '131744': 750, '131745': 750, '131746': 750,
                   '131817': 840, '131860': 612, '132006': 7800, '132025': 840, '132129': 1296, '132155': 840,
                   '132289': 1680, '132672': 1680, '132673': 1680, '132727': 840, '132728': 750, '132729': 1680,
                   '132730': 750, '132731': 1680, '132732': 840, '132733': 750, '132767': 750, '132887': 540,
                   '132921': 750, '132984': 660, '308395': 240, '308487': 288, '308563': 240, '309022': 240,
                   '309339': 180, '309581': 180, '309588': 612, '309941': 180, '434525': 750, '100557': 570,
                   '100794': 570, '100804': 570, '101046': 570, '103636': 570, '103643': 570, '103644': 570,
                   '103647': 570, '104285': 570, '108817': 570, '108818': 570, '108831': 570, '108832': 570,
                   '108833': 570, '108835': 570, '108838': 570, '111774': 570, '112635': 570, '114034': 630,
                   '114036': 570, '115679': 570, '115681': 570, '120706': 570, '123403': 570, '123408': 570,
                   '123409': 570, '125551': 570, '125857': 570, '125885': 570, '126681': 120, '126682': 120,
                   '127591': 570, '128008': 570, '128676': 570, '132007': 570, '132027': 570, '132213': 480,
                   '132249': 570, '100659': 360, '100660': 360, '100662': 720, '101198': 720, '101199': 456,
                   '101329': 2160, '102518': 240, '109754': 312, '113370': 312, '115723': 240, '123160': 240,
                   '123573': 240, '123883': 240, '124576': 240, '124864': 240, '128950': 312, '129088': 312,
                   '129294': 240, '129484': 312, '131197': 240, '131351': 264, '102361': 480, '107464': 480,
                   '107988': 396, '104301': 750, '108322': 480, '108323': 600, '108324': 6000, '113709': 432,
                   '125424': 570, '112497': 540, '112498': 600, '112499': 432, '112500': 600, '128064': 480,
                   '128066': 480, '128185': 600, '128277': 600, '132008': 600, '132009': 600, '132010': 480,
                   '112873': 600, '112912': 840, '112914': 840, '113094': 600, '113097': 360, '113172': 600,
                   '123040': 480, '112875': 750, '113757': 522, '112888': 780, '113040': 504, '120116': 504,
                   '120303': 780, '123859': 456, '124212': 888, '125225': 840, '112890': 600, '112892': 600,
                   '112895': 720, '112896': 540, '112897': 720, '112898': 540, '113043': 480, '113259': 300,
                   '113376': 1008, '113408': 1008, '117957': 648, '117958': 576, '124188': 600, '124189': 600,
                   '124447': 1008, '124448': 720, '124449': 432, '125105': 456, '126438': 432, '126583': 480,
                   '126584': 768, '126870': 648, '126908': 600, '127255': 648, '127256': 480, '127257': 960,
                   '127401': 1008, '128813': 300, '131593': 576, '131594': 480, '131595': 960, '131596': 576,
                   '131597': 480, '131598': 960, '131599': 384, '132125': 576, '310075': 960, '113862': 600,
                   '117658': 300, '117659': 240, '117684': 216, '117685': 720, '117690': 600, '117691': 1152,
                   '117692': 4320, '117693': 300, '117694': 480, '117702': 4320, '117703': 576, '117704': 600,
                   '117705': 1152, '117706': 432, '117707': 48, '117708': 360, '117709': 576, '117710': 540,
                   '117752': 300, '117753': 300, '117754': 240, '120073': 1176, '120087': 1008, '120395': 4320,
                   '120539': 384, '120550': 5760, '122204': 294, '122222': 144, '122223': 10, '124192': 432,
                   '124632': 432, '125160': 270, '125161': 270, '125162': 270, '125163': 270, '125164': 270,
                   '125165': 270, '125217': 270, '125506': 150, '125870': 18, '125890': 720, '125891': 300,
                   '125892': 360, '126864': 300, '127081': 720, '127251': 840, '127252': 360, '127551': 240,
                   '127554': 1008, '127555': 300, '127720': 150, '128231': 1584, '128444': 600, '128445': 600,
                   '128951': 210, '129295': 300, '131386': 360, '131387': 192, '131388': 300, '131389': 330,
                   '131390': 240, '131397': 360, '131400': 192, '131534': 330, '131773': 600, '131774': 480,
                   '131776': 1152, '131777': 4320, '131786': 5184, '132553': 180, '132554': 360, '132674': 360,
                   '132675': 360, '132822': 240, '132852': 864, '309019': 150, '120009': 570, '122790': 570,
                   '122795': 456, '124736': 570, '124741': 336, '124742': 288, '124743': 750, '125057': 288,
                   '125270': 570, '125539': 570, '125664': 570, '127244': 528, '130847': 288, '132146': 144,
                   '132233': 288, '132539': 144, '132540': 144, '132726': 72, '121092': 480, '125104': 480,
                   '125999': 600, '100714': 600, '103679': 432, '107331': 432, '122751': 378, '123432': 600,
                   '123434': 378, '123435': 396, '123509': 600, '123513': 378, '126435': 432, '126467': 600,
                   '126878': 180, '128278': 378, '132172': 378, '111112': 528, '123699': 528, '123766': 1110,
                   '123776': 600, '123786': 528, '123805': 570, '125788': 288, '124491': 600, '125879': 5040,
                   '126688': 216, '132156': 600, '125106': 600, '125107': 600, '125108': 600, '126290': 6000,
                   '126550': 720, '126551': 480, '126595': 576, '126596': 402, '126731': 384, '127012': 300,
                   '127887': 576, '129159': 384, '129575': 600, '131891': 576, '132026': 480, '132281': 10,
                   '132446': 576, '132453': 576, '132454': 576, '100000': 0, '122483': 780, '125133': 780,
                   '125432': 648, '125550': 720, '131666': 780, '132181': 504, '125733': 780, '125734': 576,
                   '100192': 6000, '100626': 63, '100628': 6000, '101133': 432, '101301': 576, '101415': 1080,
                   '101519': 1200, '101520': 576, '101722': 576, '102458': 720, '102459': 480, '103178': 6000,
                   '103592': 480, '103593': 42, '103921': 300, '107502': 720, '107977': 432, '110557': 480,
                   '110558': 720, '110559': 360, '110601': 672, '110990': 720, '111137': 660, '111138': 720,
                   '111460': 480, '120368': 432, '121373': 600, '121390': 4800, '122724': 720, '122920': 6000,
                   '124442': 720, '124443': 720, '124681': 384, '125228': 600, '125552': 330, '125615': 300,
                   '125629': 6000, '125856': 750, '126391': 540, '126455': 5400, '126573': 240, '126730': 2304,
                   '126757': 432, '127082': 384, '127112': 2000, '127823': 624, '127859': 720, '128722': 720,
                   '129026': 720, '130030': 330, '130783': 432, '131392': 720, '131393': 42, '131394': 432,
                   '131395': 720, '131530': 750, '131661': 1080, '131717': 1440, '131740': 360, '131894': 750,
                   '131942': 6000, '132061': 480, '132064': 6000, '132065': 720, '132068': 42, '132795': 480,
                   '132983': 660, '308471': 720, '308472': 144, '432932': 576, '100550': 420, '101628': 168,
                   '103589': 408, '103797': 3600, '104697': 144, '112295': 510, '112930': 180, '120419': 168,
                   '122755': 3600, '122780': 384, '123305': 108, '123902': 3600, '123903': 3600, '125836': 3600,
                   '126020': 144, '126144': 54, '126145': 54, '126245': 336, '126451': 3600, '126539': 144,
                   '126680': 54, '126759': 6, '127556': 144, '128287': 96, '128481': 54, '128482': 54, '129080': 108,
                   '129225': 96, '129300': 108, '129975': 96, '130845': 96, '130894': 510, '131778': 510, '131982': 120,
                   '132147': 510, '132148': 510, '132149': 510, '132186': 510, '132198': 132, '132199': 132,
                   '132333': 510, '132349': 6, '132377': 6, '132456': 132, '132480': 240, '132508': 120, '132509': 240,
                   '132519': 120, '132520': 108, '132521': 108, '132555': 480, '132582': 384, '132836': 450,
                   '132873': 8, '113372': 360, '129029': 288, '132323': 300, '132660': 360, '132661': 360,
                   '120010': 576, '122429': 1680, '124943': 384, '125540': 576, '125544': 720, '125546': 864,
                   '125549': 1152, '125789': 1152, '125790': 1152, '125791': 720, '131730': 576, '131731': 384,
                   '131732': 384, '310026': 750, '125886': 480, '125887': 480, '125888': 480, '126448': 696,
                   '126449': 696, '126450': 432, '126597': 432, '102048': 360, '121603': 720, '125244': 720,
                   '126326': 432, '126328': 1296, '127795': 144, '127933': 960, '127939': 432, '128618': 576,
                   '128643': 960, '131396': 720, '127098': 720, '112184': 3000, '111188': 432, '131198': 432,
                   '308498': 630, '131112': 432, '131113': 576, '131323': 576, '131516': 504, '132222': 504,
                   '132878': 504, '131254': 660, '131324': 540, '131806': 450, '132761': 480, '131116': 480,
                   '131217': 570, '131255': 660, '131325': 600, '132219': 480, '132225': 570, '131256': 504,
                   '131257': 432, '131258': 600, '131259': 924, '131321': 600, '131322': 924, '131818': 432,
                   '131819': 432, '131820': 924, '129165': 720, '129166': 720, '129167': 720, '129168': 720,
                   '130976': 720, '131096': 570, '131098': 570, '131099': 264, '131100': 570, '131188': 570,
                   '131260': 570, '132114': 570, '132115': 570, '132116': 570, '132117': 240, '132118': 570,
                   '132187': 576, '132393': 570, '132711': 570, '129169': 720, '129170': 720, '129171': 720,
                   '129172': 720, '129173': 720, '131101': 570, '131104': 570, '131105': 288, '131107': 570,
                   '131135': 570, '131191': 570, '131832': 570, '131833': 570, '131834': 570, '130977': 720,
                   '130978': 720, '130979': 720, '130980': 720, '131261': 600, '131262': 600, '131263': 600,
                   '131264': 600, '131183': 396, '131184': 432, '131115': 570, '131612': 570, '132049': 570,
                   '132087': 180, '132088': 180, '132089': 180, '132153': 180, '132159': 180, '132176': 180,
                   '132256': 96, '132257': 240, '132258': 240, '132259': 240, '132314': 240, '132494': 60,
                   '44300212': 180, '132240': 720, '128692': 630, '129091': 570, '129493': 570, '132470': 504,
                   '132952': 660, '132953': 576, '132954': 720, '132982': 600, '132951': 480, '132961': 4,
                   '109378': 168, '110747': 552, '130274': 312, '132342': 312, '120294': 360, '126463': 672,
                   '126801': 720, '131775': 720, '130275': 552, '131111': 456, '132265': 456, '131117': 480,
                   '132216': 480, '131118': 408, '131216': 450, '131220': 480, '132223': 480, '131231': 432,
                   '131232': 324, '131233': 480, '131340': 432, '132264': 432, '131391': 360, '132853': 360,
                   '131771': 570, '131799': 72, '132066': 720, '132461': 750, '132067': 432, '132463': 456,
                   '132127': 144, '132214': 510, '132215': 750, '132445': 750, '132227': 552, '132230': 312,
                   '132449': 4320, '132464': 360, '132857': 540, '132486': 384, '132487': 456, '310304': 570,
                   '132526': 570, '132594': 660, '132665': 570, '132667': 840, '132718': 960, '132719': 576,
                   '132720': 480, '132721': 960, '132722': 576, '132723': 0, '132785': 1680, '132850': 0, '132851': 0,
                   '132854': 300, '132855': 0, '132856': 0, '132858': 0, '132957': 480, '132958': 384, '132959': 480,
                   '132950': 720, '132917': 1680, '132955': 720, '132956': 576}


def no_mono(quant, product, quant_mono):
    if product in quant_mono:
        if quant_mono[product] == quant:
            return False
    return True


def main():
    while True:
        try:
            wb = load_workbook('Registers.RegistersBrw.xlsx')  # Загружаем файл
            break
        except FileNotFoundError:
            input('\nФайл "Registers.RegistersBrw.xlsx" не найден. '
                  'Поместите файл в одну папку с программой и нажмите Enter.')

    ws = wb['List1']  # В каком листе проверяем

    party_column = search_column(ws, 'Партия')
    q_column = search_column(ws, 'Q общ. баз.')
    cell_address_column = search_column(ws, 'Адрес ячейки')
    product_column = search_column(ws, 'Продукт')
    description_column = search_column(ws, 'Наименование')
    gtd_column = search_column(ws, 'ГТД')
    certificate_column = search_column(ws, 'Сертификат')

    dict_party = dict()  # Здесь будут все партии с их количеством

    index = 5

    party_cell = (ws[f'{party_column}{index}']).value
    q_cell = (ws[f'{q_column}{index}']).value
    cell_address_cell = (ws[f'{cell_address_column}{index}']).value
    product_cell = (ws[f'{product_column}{index}']).value
    description_cell = (ws[f'{description_column}{index}']).value
    gtd_cell = (ws[f'{gtd_column}{index}']).value
    certificate_cell = (ws[f'{certificate_column}{index}']).value

    while party_cell:
        if party_cell in dict_party:
            if no_mono(q_cell, product_cell, quant_mono_dict):
                dict_party[party_cell]['product_set'].add(product_cell)
                dict_party[party_cell]['quantity'] += 1
                dict_party[party_cell]['total'] += q_cell
                if cell_address_cell in dict_party[party_cell]['cell_addresses']['cell_address_cell']:
                    dict_party[party_cell]['cell_addresses']['cell_address_cell'] \
                        [f"{cell_address_cell}_{dict_party[party_cell]['quantity']}"] = {
                                                                                        'cell': cell_address_cell,
                                                                                        'quant_in_cell': q_cell,
                                                                                        'product': product_cell,
                                                                                        'description': description_cell,
                                                                                        'GTD': gtd_cell,
                                                                                        'certificate': certificate_cell
                                                                                         }
                else:
                    dict_party[party_cell]['cell_addresses']['cell_address_cell'][cell_address_cell] = {
                        'quant_in_cell': q_cell,
                        'cell': cell_address_cell,
                        'product': product_cell,
                        'description': description_cell,
                        'GTD': gtd_cell,
                        'certificate': certificate_cell
                    }
        else:
            if no_mono(q_cell, product_cell, quant_mono_dict):
                dict_party[party_cell] = {
                    'product_set': {product_cell},
                    'quantity': 1,
                    'total': q_cell,
                    'cell_addresses': {'cell_address_cell': {cell_address_cell: {
                        'quant_in_cell': q_cell,
                        'cell': cell_address_cell,
                        'product': product_cell,
                        'description': description_cell,
                        'GTD': gtd_cell,
                        'certificate': certificate_cell
                    }}}
                }
        index += 1
        party_cell = (ws[f'{party_column}{index}']).value
        q_cell = (ws[f'{q_column}{index}']).value
        cell_address_cell = (ws[f'{cell_address_column}{index}']).value
        product_cell = (ws[f'{product_column}{index}']).value
        description_cell = (ws[f'{description_column}{index}']).value
        gtd_cell = (ws[f'{gtd_column}{index}']).value
        certificate_cell = (ws[f'{certificate_column}{index}']).value

    wb.create_sheet('Лист1', 0)
    ws_1 = wb['Лист1']
    while True:
        try:
            max_total = int(input('Введите максимальное кол-во бутылок после объединения: '))
            break
        except ValueError:
            print('\nВведено неверное значение. Попробуйте ещё раз. ')
    # max_total = 200

    ws_1['A1'] = 'Адрес ячейки'
    ws_1['B1'] = 'Продукт'
    ws_1['C1'] = 'Наименование'
    ws_1['D1'] = 'Q общ. баз.'
    ws_1['E1'] = 'Партия'
    ws_1['F1'] = 'ГТД'
    ws_1['G1'] = 'Сертификат'

    index_new = 2
    for elem in dict_party:
        quantity = dict_party[elem]['quantity']
        total = dict_party[elem]['total']

        if quantity > 1 and total <= max_total:
            cell_address_cell = (ws_1[f'A{index_new}'])
            product_cell = (ws_1[f'B{index_new}'])
            description_cell = (ws_1[f'C{index_new}'])
            q_cell = (ws_1[f'D{index_new}'])
            party_cell = (ws_1[f'E{index_new}'])
            gtd_cell = (ws_1[f'F{index_new}'])
            certificate_cell = (ws_1[f'G{index_new}'])

            for addresses in dict_party[elem]['cell_addresses']['cell_address_cell']:
                if len(dict_party[elem]['product_set']) > 1:
                    continue
                cell_address_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses]['cell']
                q_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses]['quant_in_cell']
                product_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses]['product']
                description_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses][
                    'description']
                gtd_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses]['GTD']
                certificate_cell.value = dict_party[elem]['cell_addresses']['cell_address_cell'][addresses][
                    'certificate']
                party_cell.value = elem
                # if len(dict_party[elem]['product_set']) > 1:
                #     ws_1[f'H{index_new}'] = 'Разный товар!'
                index_new += 1
                cell_address_cell = (ws_1[f'A{index_new}'])
                product_cell = (ws_1[f'B{index_new}'])
                description_cell = (ws_1[f'C{index_new}'])
                q_cell = (ws_1[f'D{index_new}'])
                party_cell = (ws_1[f'E{index_new}'])
                gtd_cell = (ws_1[f'F{index_new}'])
                certificate_cell = (ws_1[f'G{index_new}'])
            # index_new += 1

    while True:
        try:
            wb.save('Оптимизация.xlsx')
            print('\nПрограмма завершенна. Данные сохранены в файл "Оптимизация.xlsx".')
            input('Для выхода нажмите Enter или просто закройте окно крестиком)))')
            break
        except PermissionError:
            input('\nОшибка при сохранении. '
                  'Закройте файл "Оптимизация.xlsx" и нажмите Enter.')


if __name__ == '__main__':
    main()
