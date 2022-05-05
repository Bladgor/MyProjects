# import pandas as pd
# import xlrd
# import xlwt
import openpyxl

# df = pd.DataFrame({'Name': ['Manchester City', 'Real Madrid', 'Liverpool',
#                             'FC Bayern München', 'FC Barcelona', 'Juventus'],
#                    'League': ['English Premier League (1)', 'Spain Primera Division (1)',
#                               'English Premier League (1)', 'German 1. Bundesliga (1)',
#                               'Spain Primera Division (1)', 'Italian Serie A (1)'],
#                    'TransferBudget': [176000000, 188500000, 90000000,
#                                       100000000, 180500000, 105000000]})
# df.to_excel('./teams.xlsx', sheet_name='Budgets', index=False)
#
# salaries1 = pd.DataFrame({'Name': ['L. Messi', 'Cristiano Ronaldo', 'J. Oblak'],
#                                         'Salary': [560000, 220000, 125000]})
#
# salaries2 = pd.DataFrame({'Name': ['K. De Bruyne', 'Neymar Jr', 'R. Lewandowski'],
#                                         'Salary': [370000, 270000, 240000]})
#
# salaries3 = pd.DataFrame({'Name': ['Alisson', 'M. ter Stegen', 'M. Salah'],
#                                         'Salary': [160000, 260000, 250000]})
#
# salary_sheets = {'Group1': salaries1, 'Group2': salaries2, 'Group3': salaries3}
# writer = pd.ExcelWriter('./salaries.xlsx', engine='xlsxwriter')
#
# for sheet_name in salary_sheets.keys():
#     salary_sheets[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
#
# writer.save()

# cols = [0, 1]
#
# teams = pd.read_excel('./teams.xlsx', index_col='Name', usecols=cols)
# print(teams.head())
# ========================================================================================

# открываем файл
# rb = xlrd.open_workbook('./teams.xls')
#
# # выбираем активный лист
# sheet = rb.sheet_by_index(0)

# получаем значение первой ячейки A1
# val = sheet.row_values(0)[0]
# print(val)
#
# # получаем список значений из всех записей
# values = [sheet.row_values(row_num) for row_num in range(sheet.nrows)]
# print(values)
# index = None
# for elem in values:
#     for sub_elem in elem:
#         if sub_elem == 'League':
#             index = elem.index(sub_elem)
#             break

# for elem in values:
#     print(elem[index])

# wb = xlwt.Workbook()
# ws = wb.add_sheet('Test')
#
# # в A1 записываем значение из ячейки A1 прошлого файла
# ws.write(0, 0, val[0])
#
# # в столбец B запишем нашу последовательность из столбца A исходного файла
# i = 0
# for rec in values:
#     ws.write(i, 1, rec[0])  # тут ошибка
#     i = + i

# сохраняем рабочую книгу
# wb.save('./xl_rec.xls')
# ==========================================================================================

wb = openpyxl.load_workbook(filename='./ЛЗ_май.xlsx')
sheet = wb['заказы']
print(wb.sheetnames)

# Retrieve the value of a certain cell
print(sheet['A1'].value)

# Select element 'B2' of your sheet
c = sheet['B2']

# Retrieve the row number of your element
print(f'Row = {c.row}')

# Retrieve the column letter of your element
print(f'Column = {c.column}')

# Retrieve the coordinates of the cell
print(f'Coordinate = {c.coordinate}')

# Retrieve cell value
# print(sheet.cell(row=2, column=2).value)
# index_date_calculated = 0
# for i_index in range(1, 103):
#     if sheet.cell(row=4, column=i_index).value == 'Дата-просчитан':
#         print(sheet.cell(row=4, column=i_index).value, i_index)
#         index_date_calculated = i_index
#         break

sheet['B2'] = 6
print(sheet['B2'].value)

row_4 = [x.value for x in sheet[4]]
# print(row_4)
index = row_4.index('Дата-просчитан') + 1
cell = sheet.cell(row=4, column=index)
date_calculate_column = cell.column_letter
# for elem in row_4:
#     print(elem.value, end=' ')

my_dict = {x.value: {'Дата-просчитан': str(sheet[f'{date_calculate_column}{x.row}'].value)} for x in sheet['A']}
print(my_dict)

# wb.save('ЛЗ_май.xlsx')
